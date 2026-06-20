import argparse
import json
import re
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from docx import Document


QUESTION_ID_RE = re.compile(r"Q_\d{3,}", re.IGNORECASE)
MODULE_RE = re.compile(r"^-?\s*M[oó]dulo\s*(\d+)\s*:\s*(.+)$", re.IGNORECASE)
KV_RE = re.compile(r"^\s*(?:(\d+)\.)?\s*([^:]+):\s*(.+?)\s*$")


# ============================================================
# UTILIDADES GENERALES
# ============================================================

def normalize_ws_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    if not text:
        return None
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def to_slug(text: str | None) -> str | None:
    if not text:
        return None
    t = text.lower()
    repl = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ü": "u", "ñ": "n",
        "/": "_", " ": "_", "-": "_", "(": "", ")": "", ":": "", ",": "",
        ".": "", ";": "", "[": "", "]": "", "{": "", "}": ""
    }
    for k, v in repl.items():
        t = t.replace(k, v)
    t = re.sub(r"_+", "_", t).strip("_")
    return t


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def write_json(data: dict[str, Any], path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def parse_maybe_list(value: str | None):
    if value is None:
        return None

    if isinstance(value, list):
        return value

    v = str(value).strip()

    if v.startswith("[") and v.endswith("]"):
        try:
            return json.loads(v)
        except Exception:
            pass

    return value


def coerce_bool(value: str | None):
    if value is None:
        return None

    v = normalize_ws_text(value)
    if v is None:
        return None

    low = v.lower()

    if low in {"si", "sí", "true", "yes"}:
        return True

    if low == "no" or low in {"false", "not"}:
        return False

    if "sí" in low or "si " in low or low.startswith("si"):
        return True

    if low.startswith("no"):
        return False

    return None


def normalize_answer_value(value: str | None) -> str | None:
    text = normalize_ws_text(value)
    if text is None:
        return None

    low = text.lower()

    if low in {"si", "sí"}:
        return "SI"

    if low == "no":
        return "NO"

    return text


def split_docx_lines(path: Path) -> list[str]:
    doc = Document(path)
    lines = []

    for p in doc.paragraphs:
        txt = p.text.replace("\xa0", " ").strip()
        if not txt:
            continue

        for part in re.split(r"\n+", txt):
            part = part.strip()
            if part:
                lines.append(part)

    return lines


def safe_session_id(value: str | None) -> str:
    value = value or "default"
    value = re.sub(r"[^a-zA-Z0-9_.-]", "_", value)
    return value[:120]


# ============================================================
# PARSING CATÁLOGO DE RIESGOS
# ============================================================

def detect_header_row(ws, required_labels: list[str], search_rows: int = 40) -> tuple[int, dict[str, int]]:
    best = None
    best_score = -1
    best_map = {}

    for r in range(1, min(ws.max_row, search_rows) + 1):
        row_vals = {
            c: normalize_ws_text(ws.cell(r, c).value)
            for c in range(1, ws.max_column + 1)
        }

        matches = {}
        score = 0

        for label in required_labels:
            target = label.lower()

            for c, val in row_vals.items():
                if not val:
                    continue

                low = val.lower()

                if low == target:
                    matches[label] = c
                    score += 2
                    break

                if target in low:
                    matches[label] = c
                    score += 1
                    break

        if score > best_score and len(matches) >= max(2, len(required_labels) // 2):
            best_score = score
            best = r
            best_map = matches

    if best is None:
        raise ValueError(f"No se pudo localizar la fila de cabecera con etiquetas {required_labels}")

    return best, best_map


def build_merged_lookup(ws):
    lookup = {}

    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = mr.bounds

        meta = {
            "min_col": min_col,
            "min_row": min_row,
            "max_col": max_col,
            "max_row": max_row,
            "value": normalize_ws_text(ws.cell(min_row, min_col).value),
            "is_vertical": max_row > min_row,
            "is_horizontal": max_col > min_col,
        }

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[(r, c)] = meta

    return lookup


def parse_risk_catalog(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)

    analysis_sheet = None
    analysis_header_row = None
    analysis_cols = None

    for ws in wb.worksheets:
        try:
            hr, cols = detect_header_row(
                ws,
                [
                    "ID",
                    "Preguntas del cuestionario SaaS",
                    "Control de seguridad",
                    "ID Riesgo",
                    "Nombre de riesgo",
                    "IP MEDIO",
                ],
                search_rows=15,
            )
            analysis_sheet = ws
            analysis_header_row = hr
            analysis_cols = cols
            break
        except Exception:
            pass

    if analysis_sheet is None:
        raise ValueError("No se encontró una hoja tipo 'Análisis SaaS'")

    ws = analysis_sheet

    headers = {
        c: normalize_ws_text(ws.cell(analysis_header_row, c).value)
        for c in range(1, ws.max_column + 1)
    }

    def find_col(names: list[str]):
        lower_headers = {c: (h.lower() if h else None) for c, h in headers.items()}

        for name in names:
            target = name.lower()

            for c, h in lower_headers.items():
                if h == target:
                    return c

            for c, h in lower_headers.items():
                if h and target in h:
                    return c

        return None

    c_id = find_col(["ID"])
    c_dim = find_col(["Dimensión Plataforma"])
    c_block = find_col(["Bloque"])
    c_section = find_col(["Apartado"])
    c_req = find_col(["Requisito"])
    c_q = find_col(["Preguntas del cuestionario SaaS"])
    c_ctrl = find_col(["Control de seguridad"])
    c_ctrl_id = find_col(["Control Requisitrón", "Control Requisitron", "Control Requisito"])
    c_noncomp = find_col(["Incumplimiento identificado"])
    c_assoc = find_col(["Controles asociados"])
    c_threats = find_col(["Amenazas"])
    c_risk_id = find_col(["ID Riesgo"])
    c_risk_name = find_col(["Nombre de riesgo"])
    c_risk_desc = find_col(["Descripción del riesgo"])
    c_mit = find_col(["Medidas de mitigación"])
    c_context = find_col(["Contexto (ejemplo)"])
    c_detail = find_col(["Detalle del incumplimiento (ejemplo)"])
    c_ip1 = find_col(["IP BAJO"])
    c_ip2 = find_col(["IP MEDIO"])
    c_ip3 = find_col(["IP ALTO"])
    c_mit2 = find_col(["Medidas mitigantes"])
    c_obs = find_col(["Observaciones"])

    items = []

    for r in range(analysis_header_row + 1, ws.max_row + 1):
        qid = normalize_ws_text(ws.cell(r, c_id).value) if c_id else None

        if not qid or not QUESTION_ID_RE.fullmatch(qid):
            continue

        item = {
            "question_id": qid.upper(),
            "platform_dimension": normalize_ws_text(ws.cell(r, c_dim).value) if c_dim else None,
            "block": normalize_ws_text(ws.cell(r, c_block).value) if c_block else None,
            "section": normalize_ws_text(ws.cell(r, c_section).value) if c_section else None,
            "requirement": normalize_ws_text(ws.cell(r, c_req).value) if c_req else None,
            "question_text": normalize_ws_text(ws.cell(r, c_q).value) if c_q else None,
            "control": {
                "control_id": normalize_ws_text(ws.cell(r, c_ctrl_id).value) if c_ctrl_id else None,
                "control_text": normalize_ws_text(ws.cell(r, c_ctrl).value) if c_ctrl else None,
            },
            "noncompliance_identified": normalize_ws_text(ws.cell(r, c_noncomp).value) if c_noncomp else None,
            "associated_controls": normalize_ws_text(ws.cell(r, c_assoc).value) if c_assoc else None,
            "threats": normalize_ws_text(ws.cell(r, c_threats).value) if c_threats else None,
            "risk": {
                "risk_id": normalize_ws_text(ws.cell(r, c_risk_id).value) if c_risk_id else None,
                "risk_name": normalize_ws_text(ws.cell(r, c_risk_name).value) if c_risk_name else None,
                "risk_description": normalize_ws_text(ws.cell(r, c_risk_desc).value) if c_risk_desc else None,
            },
            "mitigation_measures": normalize_ws_text(ws.cell(r, c_mit).value) if c_mit else None,
            "context_example": normalize_ws_text(ws.cell(r, c_context).value) if c_context else None,
            "noncompliance_detail_example": normalize_ws_text(ws.cell(r, c_detail).value) if c_detail else None,
            "impact_by_tier": {
                "1": normalize_ws_text(ws.cell(r, c_ip1).value) if c_ip1 else None,
                "2": normalize_ws_text(ws.cell(r, c_ip2).value) if c_ip2 else None,
                "3": normalize_ws_text(ws.cell(r, c_ip3).value) if c_ip3 else None,
            },
            "mitigation_guidance": normalize_ws_text(ws.cell(r, c_mit2).value) if c_mit2 else None,
            "observations": normalize_ws_text(ws.cell(r, c_obs).value) if c_obs else None,
            "source": {
                "sheet": ws.title,
                "row": r,
            },
        }

        items.append(item)

    risks = []

    if "Catálogo riesgos" in wb.sheetnames:
        wsr = wb["Catálogo riesgos"]

        try:
            hr, cols = detect_header_row(
                wsr,
                ["ID", "Nombre del riesgo", "Medidas de mitigación", "Criterio Aplicabilidad"],
                search_rows=10,
            )

            col_id = cols.get("ID")
            col_name = cols.get("Nombre del riesgo")
            col_mit = cols.get("Medidas de mitigación")
            col_crit = cols.get("Criterio Aplicabilidad")

            for r in range(hr + 1, wsr.max_row + 1):
                rid = normalize_ws_text(wsr.cell(r, col_id).value) if col_id else None

                if not rid or not rid.startswith("R_"):
                    continue

                risks.append(
                    {
                        "risk_id": rid,
                        "risk_name": normalize_ws_text(wsr.cell(r, col_name).value) if col_name else None,
                        "mitigation_measures": normalize_ws_text(wsr.cell(r, col_mit).value) if col_mit else None,
                        "applicability_criteria": normalize_ws_text(wsr.cell(r, col_crit).value) if col_crit else None,
                        "source": {
                            "sheet": wsr.title,
                            "row": r,
                        },
                    }
                )

        except Exception:
            pass

    threats_catalog = []

    if "Amenazas" in wb.sheetnames:
        wst = wb["Amenazas"]

        for r in range(2, wst.max_row + 1):
            val = normalize_ws_text(wst.cell(r, 1).value)

            if val:
                threats_catalog.append(
                    {
                        "threat_text": val,
                        "source": {
                            "sheet": wst.title,
                            "row": r,
                        },
                    }
                )

    return {
        "schema_version": "1.0.0",
        "artifact_type": "risk_catalog",
        "source_file": path.name,
        "source_sha256": file_sha256(path),
        "risk_catalog_version_id": f"{path.stem}-{file_sha256(path)[:8]}",
        "analysis_sheet": {
            "sheet": analysis_sheet.title,
            "header_row": analysis_header_row,
            "detected_columns": analysis_cols,
        },
        "question_mappings": items,
        "risk_catalog": risks,
        "threat_catalog": threats_catalog,
        "sheet_inventory": wb.sheetnames,
    }


def build_risk_lookup(risk_catalog: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        item["question_id"]: item
        for item in risk_catalog.get("question_mappings", [])
        if item.get("question_id")
    }


# ============================================================
# PARSING CUESTIONARIO
# ============================================================

def resolve_question_text(
    ws,
    row: int,
    col_desc: int | None,
    merged_lookup: dict,
    risk_lookup: dict[str, dict[str, Any]],
    qid: str,
):
    raw_desc = normalize_ws_text(ws.cell(row, col_desc).value) if col_desc else None
    adjacent = normalize_ws_text(ws.cell(row, col_desc + 1).value) if col_desc and (col_desc + 1) <= ws.max_column else None
    merged = merged_lookup.get((row, col_desc)) if col_desc else None

    in_vertical_merge = bool(merged and merged.get("is_vertical"))

    question_text = None
    source = None

    if in_vertical_merge and adjacent:
        question_text = adjacent
        source = "questionnaire_adjacent_column"
    elif raw_desc:
        question_text = raw_desc
        source = "questionnaire_description"
    elif adjacent:
        question_text = adjacent
        source = "questionnaire_adjacent_column"
    else:
        fallback = risk_lookup.get(qid, {}).get("question_text")
        if fallback:
            question_text = fallback
            source = "risk_catalog_fallback"

    return question_text, source


def parse_questionnaire(path: Path, risk_lookup: dict[str, dict[str, Any]] | None = None) -> dict[str, Any]:
    risk_lookup = risk_lookup or {}

    wb = openpyxl.load_workbook(path, data_only=True)

    selected = None
    max_q = -1

    for ws in wb.worksheets:
        q_count = 0

        for row in ws.iter_rows(values_only=True):
            if any(v and QUESTION_ID_RE.search(str(v)) for v in row):
                q_count += 1

        if q_count > max_q:
            max_q = q_count
            selected = ws

    if selected is None:
        raise ValueError("No se encontró una hoja de cuestionario con identificadores Q_###")

    ws = selected
    merged_lookup = build_merged_lookup(ws)

    provider_label_map = {
        "name": ["Nombre Proveedor", "Proveedor"],
        "platform_name": [
            "Servicio / Plataforma ofertada",
            "Nombre de la plataforma",
            "Plataforma",
            "Servicio ofertado",
        ],
        "url": ["URL Servicio", "URL", "URL Servicio (Producción/Demo)"],
        "functional_description": [
            "Breve descripción sobre la funcionalidad de la herramienta",
            "Funcionalidad del SaaS",
        ],
        "stores_personal_data": [
            "Indicar si se almacenan datos personales",
            "Datos personales",
        ],
    }

    provider = {k: None for k in provider_label_map}
    metadata_hits = []

    for r in range(1, min(ws.max_row, 30) + 1):
        for c in range(1, min(ws.max_column, 10) + 1):
            val = normalize_ws_text(ws.cell(r, c).value)

            if not val:
                continue

            for key, patterns in provider_label_map.items():
                if provider[key] is not None:
                    continue

                if any(p.lower() in val.lower() for p in patterns):
                    found = None

                    for cc in range(c + 1, min(ws.max_column, c + 5) + 1):
                        candidate = normalize_ws_text(ws.cell(r, cc).value)

                        if candidate:
                            found = candidate
                            break

                    provider[key] = found
                    metadata_hits.append(
                        {
                            "field": key,
                            "row": r,
                            "label": val,
                            "value": found,
                        }
                    )

    header_row, header_cols = detect_header_row(
        ws,
        ["Apartado", "ID Control", "Requisito", "Descripción", "Respuesta", "Explicación"],
        search_rows=40,
    )

    col_apartado = header_cols.get("Apartado")
    col_qid = header_cols.get("ID Control")
    col_req = header_cols.get("Requisito")
    col_desc = header_cols.get("Descripción")
    col_resp = header_cols.get("Respuesta")
    col_expl = header_cols.get("Explicación")

    header_texts = {
        c: normalize_ws_text(ws.cell(header_row, c).value)
        for c in range(1, ws.max_column + 1)
    }

    questions = []
    current_block = None
    current_subsection = None

    for r in range(header_row + 1, ws.max_row + 1):
        vals = [
            normalize_ws_text(ws.cell(r, c).value)
            for c in range(1, ws.max_column + 1)
        ]

        if not any(vals):
            continue

        apartado = normalize_ws_text(ws.cell(r, col_apartado).value) if col_apartado else None
        qid = normalize_ws_text(ws.cell(r, col_qid).value) if col_qid else None

        if apartado and not qid:
            current_block = apartado
            current_subsection = apartado
            continue

        if qid and QUESTION_ID_RE.fullmatch(qid):
            qid = qid.upper()

            if apartado:
                current_subsection = apartado
                if current_block is None:
                    current_block = apartado

            question_text, question_text_source = resolve_question_text(
                ws,
                r,
                col_desc,
                merged_lookup,
                risk_lookup,
                qid,
            )

            requirement = normalize_ws_text(ws.cell(r, col_req).value) if col_req else None

            if requirement is None:
                requirement = risk_lookup.get(qid, {}).get("requirement")
                requirement_source = "risk_catalog_fallback" if requirement else None
            else:
                requirement_source = "questionnaire"

            answer_raw = normalize_ws_text(ws.cell(r, col_resp).value) if col_resp else None
            answer_normalized = normalize_answer_value(answer_raw)
            explanation = normalize_ws_text(ws.cell(r, col_expl).value) if col_expl else None

            question = {
                "question_id": qid,
                "block": current_block,
                "section": current_subsection or current_block,
                "requirement": requirement,
                "requirement_source": requirement_source,
                "question_text": question_text,
                "question_text_source": question_text_source,
                "answer_raw": answer_raw,
                "answer": answer_normalized,
                "answer_missing": answer_normalized is None,
                "explanation": explanation,
                "source": {
                    "sheet": ws.title,
                    "row": r,
                },
            }

            parse_notes = []

            if question_text_source == "risk_catalog_fallback":
                parse_notes.append("question_text_filled_from_risk_catalog")

            if question_text_source == "questionnaire_adjacent_column":
                parse_notes.append("question_text_read_from_adjacent_column_due_to_merged_layout")

            if requirement_source == "risk_catalog_fallback":
                parse_notes.append("requirement_filled_from_risk_catalog")

            if answer_normalized != answer_raw and answer_raw is not None:
                parse_notes.append("answer_normalized")

            if answer_normalized is None:
                parse_notes.append("missing_answer_in_source")

            if parse_notes:
                question["parse_notes"] = parse_notes

            extras = {}

            for c in range(1, ws.max_column + 1):
                if c in [col_apartado, col_qid, col_req, col_desc, col_resp, col_expl]:
                    continue

                v = normalize_ws_text(ws.cell(r, c).value)
                h = header_texts.get(c)

                if v and h:
                    extras[to_slug(h)] = v

            if extras:
                question["extra_columns"] = extras

            questions.append(question)

    provider_normalized = {
        "name": provider.get("name"),
        "platform_name": provider.get("platform_name"),
        "url": provider.get("url"),
        "functional_description": provider.get("functional_description"),
        "stores_personal_data_raw": provider.get("stores_personal_data"),
        "stores_personal_data": coerce_bool(provider.get("stores_personal_data")),
    }

    return {
        "schema_version": "1.1.0",
        "artifact_type": "questionnaire",
        "source_file": path.name,
        "source_sha256": file_sha256(path),
        "questionnaire_version_id": f"{path.stem}-{file_sha256(path)[:8]}",
        "provider": provider_normalized,
        "metadata_hits": metadata_hits,
        "table_layout": {
            "sheet": ws.title,
            "header_row": header_row,
            "detected_columns": header_cols,
        },
        "normalization_policy": {
            "yes_values": ["si", "sí", "Si", "SI", "Sí"],
            "no_values": ["no", "No", "NO"],
            "normalized_closed_answers": {
                "SI": "SI",
                "NO": "NO",
            },
            "missing_answer_strategy": "keep_null_and_mark_answer_missing_true",
            "question_text_fallback_strategy": "adjacent_column_if_merged_else_risk_catalog",
        },
        "questions": questions,
    }


# ============================================================
# PARSING TRIAJE
# ============================================================

def parse_triaje(path: Path) -> dict[str, Any]:
    if path.suffix.lower() == ".docx":
        raw_lines = split_docx_lines(path)
    else:
        raw_lines = [
            ln.strip()
            for ln in path.read_text(encoding="utf-8", errors="replace").splitlines()
            if ln.strip()
        ]

    modules = []
    current = None
    fields_flat = []

    for line in raw_lines:
        m = MODULE_RE.match(line)

        if m:
            current = {
                "module_number": int(m.group(1)),
                "module_name": m.group(2).strip(),
                "fields": [],
            }
            modules.append(current)
            continue

        m = KV_RE.match(line)

        if m:
            field = {
                "item_number": int(m.group(1)) if m.group(1) else None,
                "label": m.group(2).strip(),
                "value_raw": m.group(3).strip(),
            }

            field["value"] = parse_maybe_list(field["value_raw"])

            if current is None:
                current = {
                    "module_number": None,
                    "module_name": "Sin módulo",
                    "fields": [],
                }
                modules.append(current)

            current["fields"].append(field)
            fields_flat.append(field)

    def get(*labels):
        for f in fields_flat:
            lbl = f["label"].lower()

            for l in labels:
                target = l.lower()

                if target == lbl or target in lbl:
                    return f["value"]

        return None

    personal_data_raw = get("Datos de carácter personal")

    triaje = {
        "schema_version": "1.1.0",
        "artifact_type": "triaje",
        "source_file": path.name,
        "source_sha256": file_sha256(path),
        "analysis_date": datetime.now().strftime("%d/%m/%Y"),
        "modules": modules,
        "contract": {
            "contract_name": get("Nombre de la contratación"),
            "short_description": get("Descripción breve de la contratación"),
            "providers": [get("Proveedor/es")] if get("Proveedor/es") else [],
            "business_activity": get("Características de la actividad"),
            "procurement_model": get("Gestión de la contratación"),
            "contract_duration": get("Duración del contrato"),
            "award_date_expected": get("Fecha prevista de adjudicación"),
            "language": parse_maybe_list(get("Idioma de la contratación"))
            if isinstance(get("Idioma de la contratación"), str)
            else get("Idioma de la contratación"),
            "responsible_person": get("Responsable de la contratación"),
            "project_code": get("Codigo del proyecto", "Código del proyecto"),
            "business_line": get("Línea de negocio"),
            "business_area": get("Área de negocio"),
        },
        "solution": {
            "solution_type": get("Solución tecnológica"),
            "saas_name": get("Nombre del SaaS"),
            "saas_functionality": get("Funcionalidad del SaaS"),
            "user_types": parse_maybe_list(get("Tipos de usuarios que acceden al SaaS"))
            if isinstance(get("Tipos de usuarios que acceden al SaaS"), str)
            else get("Tipos de usuarios que acceden al SaaS"),
            "connection_with_company": coerce_bool(str(get("Conexión del SaaS con EMPRESA")))
            if get("Conexión del SaaS con EMPRESA") is not None
            else None,
            "involves_ai": coerce_bool(str(get("La actividad conlleva IA")))
            if get("La actividad conlleva IA") is not None
            else None,
        },
        "information": {
            "classification": get("Tipo de información"),
            "classification_detail": get("Detalle del tipo de información"),
            "personal_data": True
            if "sí" in str(personal_data_raw).lower()
            else (False if personal_data_raw else None),
            "personal_data_regime": "GDPR"
            if "gdpr" in str(personal_data_raw).lower()
            else None,
            "personal_data_types": parse_maybe_list(get("Tipos de datos de carácter personal"))
            if isinstance(get("Tipos de datos de carácter personal"), str)
            else get("Tipos de datos de carácter personal"),
            "personal_data_volume": get("Volumen de datos personales que se va a tratar"),
            "processing_location": get("Lugar donde se realiza el tratamiento de la información"),
        },
        "provider_interaction": {
            "provider_access_to_company_infrastructure": False
            if "no acceden" in str(get("Número de usuarios del proveedor que acceden a EMPRESA")).lower()
            else None,
            "provider_access_detail": get("Número de usuarios del proveedor que acceden a EMPRESA"),
            "systems_interaction": get("Con qué sistemas/plataformas de EMPRESA interactua el proveedor"),
        },
        "risk_context": {
            "service_is_critical": coerce_bool(str(get("El servicio, proyecto o solución es critico")))
            if get("El servicio, proyecto o solución es critico") is not None
            else None,
        },
    }

    triaje["derived_flags"] = {
        "has_personal_data": triaje["information"]["personal_data"],
        "is_saas": "saas" in str(triaje["solution"]["solution_type"]).lower()
        if triaje["solution"]["solution_type"]
        else None,
        "has_provider_internal_access": triaje["provider_interaction"]["provider_access_to_company_infrastructure"],
        "involves_ai": triaje["solution"]["involves_ai"],
    }

    return triaje


# ============================================================
# CASE META / CHECK COMPLETE
# ============================================================

def get_case_summary(triaje: dict[str, Any], questionnaire: dict[str, Any]) -> dict[str, Any]:
    return {
        "project_code": triaje.get("contract", {}).get("project_code"),
        "business_line": triaje.get("contract", {}).get("business_line"),
        "business_area": triaje.get("contract", {}).get("business_area"),
        "provider": triaje.get("contract", {}).get("providers"),
        "contract_name": triaje.get("contract", {}).get("contract_name"),
        "solution_type": triaje.get("solution", {}).get("solution_type"),
        "saas_name": triaje.get("solution", {}).get("saas_name"),
        "classification": triaje.get("information", {}).get("classification"),
        "personal_data": triaje.get("information", {}).get("personal_data"),
        "personal_data_regime": triaje.get("information", {}).get("personal_data_regime"),
        "questionnaire_questions": len(questionnaire.get("questions", [])),
        "questions_with_missing_answer": sum(
            1 for item in questionnaire.get("questions", [])
            if item.get("answer_missing")
        ),
    }


def build_case_meta(
    case_id: str,
    case_id_raw: str | None,
    tier: int,
    jsondir: Path,
    created_by: str = "n8n",
) -> dict[str, Any]:
    triaje_path = jsondir / "triaje.json"
    questionnaire_path = jsondir / "questionnaire.json"
    case_meta_path = jsondir / "case_meta.json"

    if not triaje_path.exists():
        raise FileNotFoundError(f"No existe triaje.json en {jsondir}")

    if not questionnaire_path.exists():
        raise FileNotFoundError(f"No existe questionnaire.json en {jsondir}")

    triaje = read_json(triaje_path)
    questionnaire = read_json(questionnaire_path)

    if not case_id_raw:
        case_id_raw = (
            triaje.get("case", {}).get("case_id_raw")
            or triaje.get("contract", {}).get("project_code")
            or case_id
        )

    case_meta = {
        "schema_version": "1.0.0",
        "artifact_type": "case_meta",
        "case_id": case_id,
        "case_id_raw": case_id_raw,
        "tier": int(tier),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "created_by": created_by,
        "status": "ready_for_risk_analysis",
        "inputs": {
            "triaje_json": str(triaje_path),
            "questionnaire_json": str(questionnaire_path),
        },
        "summary": get_case_summary(triaje, questionnaire),
    }

    write_json(case_meta, case_meta_path)

    return case_meta


def check_case_complete(jsondir: Path) -> dict[str, Any]:
    triaje_path = jsondir / "triaje.json"
    questionnaire_path = jsondir / "questionnaire.json"
    case_meta_path = jsondir / "case_meta.json"

    triaje_exists = triaje_path.exists()
    questionnaire_exists = questionnaire_path.exists()
    case_meta_exists = case_meta_path.exists()

    complete_without_tier = triaje_exists and questionnaire_exists
    complete_with_tier = complete_without_tier and case_meta_exists

    if complete_with_tier:
        status = "COMPLETE_WITH_TIER"
    elif complete_without_tier:
        status = "COMPLETE_NEEDS_TIER"
    else:
        status = "PENDING"

    return {
        "ok": True,
        "command": "check-complete",
        "status": status,
        "complete_without_tier": complete_without_tier,
        "complete_with_tier": complete_with_tier,
        "files": {
            "triaje_json": str(triaje_path),
            "triaje_exists": triaje_exists,
            "questionnaire_json": str(questionnaire_path),
            "questionnaire_exists": questionnaire_exists,
            "case_meta_json": str(case_meta_path),
            "case_meta_exists": case_meta_exists,
        },
    }


# ============================================================
# CHAT TIER
# ============================================================

def find_cases_needing_tier(cases_root: Path) -> list[dict[str, Any]]:
    pending = []

    if not cases_root.exists():
        return pending

    for case_dir in sorted(cases_root.iterdir()):
        if not case_dir.is_dir():
            continue

        json_dir = case_dir / "02_json"
        triaje_path = json_dir / "triaje.json"
        questionnaire_path = json_dir / "questionnaire.json"
        case_meta_path = json_dir / "case_meta.json"

        if triaje_path.exists() and questionnaire_path.exists() and not case_meta_path.exists():
            case_id = case_dir.name
            case_id_raw = case_id

            try:
                triaje = read_json(triaje_path)
                case_id_raw = (
                    triaje.get("case", {}).get("case_id_raw")
                    or triaje.get("contract", {}).get("project_code")
                    or case_id
                )
            except Exception:
                pass

            pending.append(
                {
                    "case_id": case_id,
                    "case_id_raw": case_id_raw,
                    "json_dir": str(json_dir),
                }
            )

    return pending


def extract_tier_from_message(message: str | None) -> int | None:
    if not message:
        return None

    patterns = [
        r"\bTIER\s*[:=]?\s*([123])\b",
        r"\btier\s*[:=]?\s*([123])\b",
        r"\b([123])\b",
    ]

    for pattern in patterns:
        m = re.search(pattern, message, re.IGNORECASE)

        if m:
            return int(m.group(1))

    return None


def match_case_from_message(message: str, pending: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not message:
        return None

    low = message.lower()

    for item in pending:
        cid = str(item.get("case_id", "")).lower()
        cid_raw = str(item.get("case_id_raw", "")).lower()

        if cid and cid in low:
            return item

        if cid_raw and cid_raw in low:
            return item

    return None


def handle_chat_tier(message: str, session_id: str, cases_root: Path) -> dict[str, Any]:
    state_dir = cases_root.parent / "chat_state"
    state_dir.mkdir(parents=True, exist_ok=True)

    session_path = state_dir / f"{safe_session_id(session_id)}.json"

    state = {}

    if session_path.exists():
        try:
            state = read_json(session_path)
        except Exception:
            state = {}

    pending = find_cases_needing_tier(cases_root)
    tier = extract_tier_from_message(message)
    selected_case = state.get("selected_case")

    # Caso 1: ya había un caso seleccionado y el usuario responde con TIER
    if tier and selected_case:
        case_id = selected_case["case_id"]
        case_id_raw = selected_case.get("case_id_raw", case_id)
        jsondir = Path(selected_case["json_dir"])

        case_meta = build_case_meta(
            case_id=case_id,
            case_id_raw=case_id_raw,
            tier=tier,
            jsondir=jsondir,
            created_by="n8n_chat_agent",
        )

        state = {
            "selected_case": None,
            "last_completed_case": {
                "case_id": case_id,
                "case_id_raw": case_id_raw,
                "tier": tier,
            },
        }

        write_json(state, session_path)

        return {
            "ok": True,
            "action": "tier_saved",
            "case_id": case_id,
            "case_id_raw": case_id_raw,
            "tier": tier,
            "message": (
                f"Perfecto. He guardado el TIER {tier} para el caso {case_id_raw}. "
                f"El caso queda listo para iniciar el análisis de riesgos."
            ),
            "case_meta": case_meta,
        }

    # Caso 2: no hay casos pendientes
    if not pending:
        return {
            "ok": True,
            "action": "no_pending_cases",
            "message": (
                "No hay casos pendientes de TIER. "
                "Cuando se reciban el triaje y el cuestionario, podré solicitarlo aquí."
            ),
        }

    # Caso 3: el usuario dice caso + TIER en el mismo mensaje
    matched_case = match_case_from_message(message, pending)

    if tier and matched_case:
        case_id = matched_case["case_id"]
        case_id_raw = matched_case.get("case_id_raw", case_id)
        jsondir = Path(matched_case["json_dir"])

        case_meta = build_case_meta(
            case_id=case_id,
            case_id_raw=case_id_raw,
            tier=tier,
            jsondir=jsondir,
            created_by="n8n_chat_agent",
        )

        return {
            "ok": True,
            "action": "tier_saved",
            "case_id": case_id,
            "case_id_raw": case_id_raw,
            "tier": tier,
            "message": (
                f"Perfecto. He guardado el TIER {tier} para el caso {case_id_raw}. "
                f"El caso queda listo para iniciar el análisis de riesgos."
            ),
            "case_meta": case_meta,
        }

    # Caso 4: solo hay un caso pendiente. Lo seleccionamos y preguntamos TIER
    if len(pending) == 1:
        selected = pending[0]

        state["selected_case"] = selected
        write_json(state, session_path)

        return {
            "ok": True,
            "action": "ask_tier",
            "case_id": selected["case_id"],
            "case_id_raw": selected["case_id_raw"],
            "message": (
                f"El caso {selected['case_id_raw']} ya tiene triaje y cuestionario, "
                f"pero todavía falta indicar el TIER.\n\n"
                f"¿Cuál es el TIER del caso?\n\n"
                f"Responde con uno de estos formatos:\n"
                f"TIER: 1\n"
                f"TIER: 2\n"
                f"TIER: 3"
            ),
        }

    # Caso 5: hay varios casos pendientes
    cases_text = "\n".join(
        f"- {item['case_id_raw']} ({item['case_id']})"
        for item in pending
    )

    return {
        "ok": True,
        "action": "multiple_pending_cases",
        "message": (
            "Hay varios casos pendientes de TIER:\n\n"
            f"{cases_text}\n\n"
            "Indica el caso y el TIER. Ejemplo:\n"
            "Codigo_del_proyecto TIER: 2"
        ),
        "pending_cases": pending,
    }


# ============================================================
# COMANDOS
# ============================================================

def command_parse_triaje(args):
    input_path = Path(args.input)
    outdir = Path(args.outdir)

    if not input_path.exists():
        raise FileNotFoundError(f"No existe el archivo de triaje: {input_path}")

    triaje = parse_triaje(input_path)

    triaje["case"] = {
        "case_id": args.case_id,
        "case_id_raw": args.case_id_raw or args.case_id,
        "source_subject": args.subject,
        "ingested_at": datetime.now().isoformat(timespec="seconds"),
    }

    write_json(triaje, outdir / "triaje.json")

    summary = {
        "ok": True,
        "command": "parse-triaje",
        "case_id": args.case_id,
        "case_id_raw": args.case_id_raw or args.case_id,
        "output": str(outdir / "triaje.json"),
        "triaje_modules": len(triaje.get("modules", [])),
        "fields_count": sum(len(m.get("fields", [])) for m in triaje.get("modules", [])),
        "contract": triaje.get("contract", {}),
        "solution": triaje.get("solution", {}),
        "information": triaje.get("information", {}),
    }

    write_json(summary, outdir / "parse_summary_triaje.json")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def command_parse_questionnaire(args):
    input_path = Path(args.input)
    outdir = Path(args.outdir)

    if not input_path.exists():
        raise FileNotFoundError(f"No existe el cuestionario: {input_path}")

    risk_lookup = {}

    if args.risk_workbook:
        risk_workbook_path = Path(args.risk_workbook)

        if risk_workbook_path.exists():
            rc = parse_risk_catalog(risk_workbook_path)
            risk_lookup = build_risk_lookup(rc)

            if args.write_risk_catalog:
                write_json(rc, outdir / "risk_catalog.json")

    q = parse_questionnaire(input_path, risk_lookup=risk_lookup)

    q["case"] = {
        "case_id": args.case_id,
        "ingested_at": datetime.now().isoformat(timespec="seconds"),
    }

    write_json(q, outdir / "questionnaire.json")

    summary = {
        "ok": True,
        "command": "parse-questionnaire",
        "case_id": args.case_id,
        "output": str(outdir / "questionnaire.json"),
        "questionnaire_questions": len(q.get("questions", [])),
        "questions_with_missing_answer": sum(
            1 for item in q.get("questions", [])
            if item.get("answer_missing")
        ),
        "questions_filled_from_risk_catalog": sum(
            1 for item in q.get("questions", [])
            if item.get("question_text_source") == "risk_catalog_fallback"
        ),
        "questions_read_from_adjacent_column": sum(
            1 for item in q.get("questions", [])
            if item.get("question_text_source") == "questionnaire_adjacent_column"
        ),
        "provider": q.get("provider", {}),
    }

    write_json(summary, outdir / "parse_summary_questionnaire.json")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def command_parse_risk_catalog(args):
    input_path = Path(args.input)
    outdir = Path(args.outdir)

    if not input_path.exists():
        raise FileNotFoundError(f"No existe el catálogo de riesgos: {input_path}")

    rc = parse_risk_catalog(input_path)
    write_json(rc, outdir / "risk_catalog.json")

    summary = {
        "ok": True,
        "command": "parse-risk-catalog",
        "output": str(outdir / "risk_catalog.json"),
        "risk_question_mappings": len(rc.get("question_mappings", [])),
        "risk_catalog_items": len(rc.get("risk_catalog", [])),
        "threat_catalog_items": len(rc.get("threat_catalog", [])),
        "sheet_inventory": rc.get("sheet_inventory", []),
    }

    write_json(summary, outdir / "parse_summary_risk_catalog.json")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def command_build_case(args):
    jsondir = Path(args.jsondir)

    if args.tier not in [1, 2, 3]:
        raise ValueError("El TIER debe ser 1, 2 o 3")

    case_meta = build_case_meta(
        case_id=args.case_id,
        case_id_raw=args.case_id_raw or args.case_id,
        tier=args.tier,
        jsondir=jsondir,
        created_by="n8n",
    )

    result = {
        "ok": True,
        "command": "build-case",
        "case_meta_json": str(jsondir / "case_meta.json"),
        "case_meta": case_meta,
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))


def command_check_complete(args):
    result = check_case_complete(Path(args.jsondir))
    print(json.dumps(result, ensure_ascii=False, indent=2))


def command_parse_all(args):
    questionnaire_path = Path(args.questionnaire)
    risk_workbook_path = Path(args.risk_workbook)
    triaje_path = Path(args.triaje)
    outdir = Path(args.outdir)

    for label, path in [
        ("questionnaire", questionnaire_path),
        ("risk-workbook", risk_workbook_path),
        ("triaje", triaje_path),
    ]:
        if not path.exists():
            raise FileNotFoundError(f"No se encontró el archivo para --{label}: {path}")

    rc = parse_risk_catalog(risk_workbook_path)
    risk_lookup = build_risk_lookup(rc)
    q = parse_questionnaire(questionnaire_path, risk_lookup=risk_lookup)
    tr = parse_triaje(triaje_path)

    write_json(q, outdir / "questionnaire.json")
    write_json(rc, outdir / "risk_catalog.json")
    write_json(tr, outdir / "triaje.json")

    summary = {
        "ok": True,
        "command": "parse-all",
        "questionnaire_questions": len(q.get("questions", [])),
        "risk_question_mappings": len(rc.get("question_mappings", [])),
        "risk_catalog_items": len(rc.get("risk_catalog", [])),
        "threat_catalog_items": len(rc.get("threat_catalog", [])),
        "triaje_modules": len(tr.get("modules", [])),
        "questions_with_missing_answer": sum(
            1 for item in q.get("questions", [])
            if item.get("answer_missing")
        ),
        "questions_filled_from_risk_catalog": sum(
            1 for item in q.get("questions", [])
            if item.get("question_text_source") == "risk_catalog_fallback"
        ),
        "questions_read_from_adjacent_column": sum(
            1 for item in q.get("questions", [])
            if item.get("question_text_source") == "questionnaire_adjacent_column"
        ),
    }

    write_json(summary, outdir / "parse_summary.json")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def command_chat_tier(args):
    result = handle_chat_tier(
        message=args.message,
        session_id=args.session_id,
        cases_root=Path(args.cases_root),
    )

    print(json.dumps(result, ensure_ascii=False, indent=2))


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Parser de ingesta para n8n: triaje, cuestionario, catálogo y gestión de TIER por chat"
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    # parse-triaje
    p_triaje = subparsers.add_parser(
        "parse-triaje",
        help="Parsea el triaje desde un TXT/DOCX y genera triaje.json",
    )
    p_triaje.add_argument("--input", required=True, type=Path)
    p_triaje.add_argument("--outdir", required=True, type=Path)
    p_triaje.add_argument("--case-id", required=True)
    p_triaje.add_argument("--case-id-raw", required=False)
    p_triaje.add_argument("--subject", required=False)
    p_triaje.set_defaults(func=command_parse_triaje)

    # parse-questionnaire
    p_questionnaire = subparsers.add_parser(
        "parse-questionnaire",
        help="Parsea el cuestionario XLSX y genera questionnaire.json",
    )
    p_questionnaire.add_argument("--input", required=True, type=Path)
    p_questionnaire.add_argument("--outdir", required=True, type=Path)
    p_questionnaire.add_argument("--case-id", required=True)
    p_questionnaire.add_argument("--risk-workbook", required=False, type=Path)
    p_questionnaire.add_argument("--write-risk-catalog", action="store_true")
    p_questionnaire.set_defaults(func=command_parse_questionnaire)

    # parse-risk-catalog
    p_risk = subparsers.add_parser(
        "parse-risk-catalog",
        help="Parsea el Excel del catálogo de riesgos y genera risk_catalog.json",
    )
    p_risk.add_argument("--input", required=True, type=Path)
    p_risk.add_argument("--outdir", required=True, type=Path)
    p_risk.set_defaults(func=command_parse_risk_catalog)

    # build-case
    p_build = subparsers.add_parser(
        "build-case",
        help="Genera case_meta.json cuando ya existen triaje.json y questionnaire.json",
    )
    p_build.add_argument("--case-id", required=True)
    p_build.add_argument("--case-id-raw", required=False)
    p_build.add_argument("--tier", required=True, type=int, choices=[1, 2, 3])
    p_build.add_argument("--jsondir", required=True, type=Path)
    p_build.set_defaults(func=command_build_case)

    # check-complete
    p_check = subparsers.add_parser(
        "check-complete",
        help="Comprueba si un caso tiene triaje.json, questionnaire.json y case_meta.json",
    )
    p_check.add_argument("--jsondir", required=True, type=Path)
    p_check.set_defaults(func=command_check_complete)

    # chat-tier
    p_chat = subparsers.add_parser(
        "chat-tier",
        help="Gestiona por chat la solicitud y guardado del TIER",
    )
    p_chat.add_argument("--message", required=True)
    p_chat.add_argument("--session-id", required=False, default="default")
    p_chat.add_argument("--cases-root", required=True, type=Path)
    p_chat.set_defaults(func=command_chat_tier)

    # parse-all
    p_all = subparsers.add_parser(
        "parse-all",
        help="Modo completo antiguo: parsea cuestionario, catálogo y triaje en una sola ejecución",
    )
    p_all.add_argument("--questionnaire", required=True, type=Path)
    p_all.add_argument("--risk-workbook", required=True, type=Path)
    p_all.add_argument("--triaje", required=True, type=Path)
    p_all.add_argument("--outdir", required=True, type=Path)
    p_all.set_defaults(func=command_parse_all)

    args = parser.parse_args()

    try:
        args.func(args)
    except Exception as exc:
        error = {
            "ok": False,
            "command": args.command,
            "error_type": type(exc).__name__,
            "error": str(exc),
        }
        print(json.dumps(error, ensure_ascii=False, indent=2))
        raise


if __name__ == "__main__":
    main()